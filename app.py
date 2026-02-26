import os
import shutil
import json
import sqlite3
import re
import csv
import io
import secrets
import bcrypt
from datetime import datetime
from flask import Flask, request, jsonify, send_file, send_from_directory, render_template, session, redirect, url_for
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from apscheduler.schedulers.background import BackgroundScheduler

# Carregar arquivo .env se existir
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv não instalado, usar variáveis de sistema

app = Flask(__name__, static_folder='static', template_folder='templates')
# Gerar secret_key segura de forma dinâmica
app.secret_key = os.environ.get('SECRET_KEY') or secrets.token_hex(32)

# rate limiting
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# Agendador de tarefas
scheduler = BackgroundScheduler()
scheduler.start()

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(BASE_DIR, 'cod_tasy.json')
DB_PATH = os.path.join(BASE_DIR, 'laudos.db')
PDF_STORAGE = os.path.join(BASE_DIR, 'static', 'laudos')
LEGACY_DIR = os.path.join(BASE_DIR, 'Laudos antigos')
NETWORK_PATH_DEFAULT = r'\\10.0.2.35\Operadora\Departamentos\TI\Backup Laudos 2026' # Default if not in DB

if not os.path.exists(PDF_STORAGE):
    os.makedirs(PDF_STORAGE)

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Table for users
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'suporte',
            full_name TEXT,
            is_active INTEGER DEFAULT 1,
            requires_password_change INTEGER DEFAULT 0
        )
    ''')
    
    cursor.execute("PRAGMA table_info(users)")
    user_columns = [row[1] for row in cursor.fetchall()]
    user_new_cols = [
        ('role', 'TEXT DEFAULT "suporte"'),
        ('full_name', 'TEXT'),
        ('is_active', 'INTEGER DEFAULT 1'),
        ('requires_password_change', 'INTEGER DEFAULT 0')
    ]
    for col_name, col_type in user_new_cols:
        if col_name not in user_columns:
            # Validar nomes de coluna para prevenir SQL injection
            if col_name in ['role', 'full_name', 'is_active', 'requires_password_change']:
                cursor.execute(f"ALTER TABLE users ADD COLUMN {col_name} {col_type}")
    # Table for laudos
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS laudos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            id_laudo TEXT NOT NULL,
            data TEXT NOT NULL,
            unidade TEXT,
            setor TEXT,
            local TEXT,
            nome_analista TEXT,
            descricao_problema TEXT,
            marca TEXT,
            modelo TEXT,
            serie TEXT,
            situacao TEXT,
            item_defeito TEXT,
            cargo_analista TEXT,
            itens_count INTEGER,
            is_test INTEGER DEFAULT 0,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cursor.execute("PRAGMA table_info(laudos)")
    columns = [row[1] for row in cursor.fetchall()]
    new_cols = [
        ('marca', 'TEXT'), 
        ('modelo', 'TEXT'), 
        ('serie', 'TEXT'), 
        ('situacao', 'TEXT'),
        ('item_defeito', 'TEXT'),
        ('cargo_analista', 'TEXT'),
        ('is_test', 'INTEGER DEFAULT 0'),
        ('tipo', 'TEXT DEFAULT "laudo"'),
        ('chamado', 'TEXT')
    ]
    for col_name, col_type in new_cols:
        if col_name not in columns:
            # Validar nomes de coluna para prevenir SQL injection
            valid_cols = ['marca', 'modelo', 'serie', 'situacao', 'item_defeito', 'cargo_analista', 'is_test', 'tipo', 'chamado']
            if col_name in valid_cols:
                cursor.execute(f"ALTER TABLE laudos ADD COLUMN {col_name} {col_type}")
    cursor.execute('SELECT password FROM users WHERE username = "admin"')
    existing_admin = cursor.fetchone()
    # Hash padrão segura para admin inicial
    DEFAULT_ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'ChangeMe@123')
    default_password_hash = bcrypt.hashpw(DEFAULT_ADMIN_PASSWORD.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    if not existing_admin:
        cursor.execute('INSERT INTO users (username, password, role, full_name) VALUES (?, ?, ?, ?)', 
                      ('admin', default_password_hash, 'master', 'Administrador'))
    else:
        # Apenas atualize se ainda estiver com a senha padrão antiga (para migração)
        print("⚠️  Admin já existe. Se necessário redefinir, use o painel de administração.")
    
    # Table for settings
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    # Table for options
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS options (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT NOT NULL,
            value TEXT NOT NULL,
            extra TEXT
        )
    ''')
    
    # Initialize default network path if not exists
    cursor.execute('SELECT * FROM settings WHERE key = "network_path"')
    if not cursor.fetchone():
        cursor.execute('INSERT INTO settings (key, value) VALUES ("network_path", ?)', (NETWORK_PATH_DEFAULT,))

    # Seed default options if empty
    cursor.execute('SELECT COUNT(*) FROM options')
    if cursor.fetchone()[0] == 0:
        defaults = {
            "unidades": [
                "Operadora", "CG24h", "CG Ambulatório", "Centro Pediátrico",
                "Medicina Preventiva", "WorkMed", "Centro Médico Bangu",
                "Centro Médico Sulacap", "Vila Nova", "Centro Social Bangu",
                "Quality Gold", "CETI", "CT - Areia Branca", "ITCM",
                "CM Seropédica", "HGSC", "Bangu"
            ],
            "setores": [
                "Administrativo", "Atendimento", "Auditoria", "Cadastro",
                "Call Center", "Cobrança", "Comercial PJ", "Compras",
                "Contabilidade", "Contas Médicas", "Controladoria",
                "Credenciamento", "DEREG - OPME - Auditoria", "Diretoria",
                "Enfermagem", "Faturamento", "Fidelização", "Financeiro",
                "Juridico", "Manutenção", "Patrimônio", "Presidência",
                "Recuperação de Produtos", "Recursos Humanos",
                "Relacionamento Empresarial", "TI"
            ],
            "locais": ["Campo Grande", "Bangu", "Sulacap", "Santa Cruz", "Itaguaí", "Seropédica"],
            "analistas": [
                {"nome": "Marcus Vinicius de Oliveira", "cargo": "Analista de TI"},
                {"nome": "Tulio Maravilha", "cargo": "Analista de TI"},
                {"nome": "Kevin Vanucci", "cargo": "Gestor de TI"},
                {"nome": "Jonathan Villas Boas", "cargo": "Assistente de TI"},
                {"nome": "Romayne de Arruda Matos", "cargo": "Analista de TI"}
            ],
            "marcas": ["Dell", "HP", "Lenovo", "Unifi", "Jabra", "Ubiquiti", "Logitech", "Kingston", "Avaya", "Samsung", "LG"],
            "modelos": ["OPTIPLEX 3070", "OPTIPLEX 3020", "PRODESK 400 G5", "Latitude 3420", "MK120", "A400 SATA III", "U7-pro", "Precison 3551", "Vostro 3470", "G3 3500"],
            "itens": ["Placa mãe", "Processador", "Memória", "Disco (HD/SSD)", "Fonte", "Placa de Rede", "Monitor", "Teclado", "Mouse", "Nobreak", "Estabilizador", "Switch", "Telefone", "Impressora", "Câmera", "Office", "Acess Point / Unifi", "Adaptadores", "Headsets", "Computador", "SSD"],
            "cargos": ["Assistente de TI", "Analista de TI", "Gestor de TI", "Gerente de TI", "Técnico Responsável", "Coordenador de Infraestrutura"]
        }
        for cat, vals in defaults.items():
            for v in vals:
                # v can be a string or a dict (for analistas)
                val = v.get('nome') if isinstance(v, dict) else v
                
                # Check if item already exists in this category
                cursor.execute('SELECT 1 FROM options WHERE category = ? AND value = ?', (cat, val))
                if not cursor.fetchone():
                    if isinstance(v, dict):
                        # For analistas
                        cursor.execute('INSERT INTO options (category, value, extra) VALUES (?, ?, ?)',
                                       (cat, val, json.dumps({'cargo': v.get('cargo')})))
                    else:
                        cursor.execute('INSERT INTO options (category, value) VALUES (?, ?)', (cat, v))

    conn.commit()
    conn.close()

init_db()

def load_tasy_data():
    """Load all tasy codes from the JSON file."""
    if not os.path.exists(JSON_PATH):
        print(f"Aviso: {JSON_PATH} não encontrado.")
        return []
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

# Pre-load data on startup
try:
    TASY_DATA_CACHE = load_tasy_data()
except Exception:
    TASY_DATA_CACHE = []

@app.route('/')
def index():
    if 'user_id' in session and session.get('role') == 'viewer':
        return redirect(url_for('viewer'))
    return render_template('index.html')

@app.route('/viewer')
def viewer():
    if 'user_id' not in session or session.get('role') != 'viewer':
        return redirect(url_for('index'))
    return render_template('viewer.html')

@app.route('/config')
def config():
    if 'user_id' not in session:
        return redirect(url_for('index'))
    return render_template('config.html')

@app.after_request
def add_header(response):
    if request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

from functools import wraps

def role_required(roles):
    """Decorator to restrict access to specific roles."""
    if isinstance(roles, str):
        roles = [roles]
        
    def decorator(f):
       @wraps(f)
       def decorated_function(*args, **kwargs):
           if 'user_id' not in session:
               return jsonify({'success': False, 'error': 'Não autorizado'}), 401
           
           if session.get('role') not in roles and 'master' not in roles: # master usually bypasses unless specific
               # If role is master, allow everything if master is in allowed roles or if we want global access
               if session.get('role') != 'master':
                   return jsonify({'success': False, 'error': 'Acesso negado: permissão insuficiente'}), 403
           
           # If master exists in session, it passes if master is in roles OR if we just allow master everywhere
           if session.get('role') == 'master':
               return f(*args, **kwargs)
               
           if session.get('role') in roles:
               return f(*args, **kwargs)
               
           return jsonify({'success': False, 'error': 'Acesso negado: permissão insuficiente'}), 403
       return decorated_function
    return decorator

# ============================================================
# FUNÇÕES DE BACKUP E AGENDAMENTO
# ============================================================

def perform_backup():
    """Executa o backup dos laudos para a rede (função reutilizável)"""
    import os
    import shutil
    from datetime import datetime
    
    try:
        # Obter caminho de rede configurado
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT value FROM settings WHERE key = "network_path"')
        row = cursor.fetchone()
        conn.close()
        
        network_path = row[0] if row else NETWORK_PATH_DEFAULT
        
        if not network_path or not os.path.exists(network_path):
            print(f'[BACKUP] Caminho de rede não acessível: {network_path}')
            return False, f'Caminho não acessível: {network_path}', 0
        
        # Criar subpasta com data
        backup_date = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        backup_folder = os.path.join(network_path, f'backup_{backup_date}')
        os.makedirs(backup_folder, exist_ok=True)
        
        # Copiar PDFs da pasta static/laudos
        laudos_source = 'static/laudos'
        files_count = 0
        
        if os.path.exists(laudos_source):
            for file in os.listdir(laudos_source):
                if file.endswith('.pdf'):
                    src = os.path.join(laudos_source, file)
                    dst = os.path.join(backup_folder, file)
                    shutil.copy2(src, dst)
                    files_count += 1
        
        # Salvar data do último backup
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)',
                      ('last_backup', datetime.now().isoformat()))
        conn.commit()
        conn.close()
        
        print(f'[BACKUP] Concluído: {files_count} arquivo(s) em {backup_folder}')
        return True, f'{files_count} arquivo(s) copiado(s)', files_count
        
    except Exception as e:
        print(f'[BACKUP] Erro: {str(e)}')
        return False, f'Erro ao fazer backup: {str(e)}', 0

def setup_backup_scheduler():
    """Configura o agendador de backup diário"""
    try:
        # Obter configuração do banco de dados
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT value FROM settings WHERE key = "auto_backup_enabled"')
        row = cursor.fetchone()
        auto_backup = row[0] == '1' if row else False
        conn.close()
        
        # Remover job anterior se existir
        if scheduler.get_job('daily_backup'):
            scheduler.remove_job('daily_backup')
        
        # Agendar se habilitado
        if auto_backup:
            scheduler.add_job(
                perform_backup,
                'cron',
                hour=20,
                minute=0,
                id='daily_backup',
                name='Daily Backup at 20:00'
            )
            print('[SCHEDULER] Backup diário agendado para 20:00')
        else:
            print('[SCHEDULER] Backup automático desabilitado')
            
    except Exception as e:
        print(f'[SCHEDULER] Erro ao configurar: {str(e)}')

# Configurar scheduler na inicialização
setup_backup_scheduler()

@app.route('/api/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")  # Rate limit for login
def login():
    if request.method == 'GET':
        if 'user_id' in session:
            return jsonify({
                'success': True,
                'username': session.get('username'),
                'role': session.get('role'),
                'full_name': session.get('full_name'),
                'requires_password_change': session.get('requires_password_change', False)
            })
        return jsonify({'success': False, 'error': 'Não logado'}), 401

    data = request.get_json()
    user = data.get('username')
    pw = data.get('password')
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, password, role, full_name, is_active, requires_password_change FROM users WHERE username = ?', (user,))
    row = cursor.fetchone()
    conn.close()
    
    if row and bcrypt.checkpw(pw.encode('utf-8'), row[2].encode('utf-8')):
        if not row[5]: # is_active
            return jsonify({'success': False, 'error': 'Usuário inativo'}), 403
            
        session['user_id'] = row[0]
        session['username'] = row[1]
        session['role'] = row[3]
        session['full_name'] = row[4]
        session['requires_password_change'] = bool(row[6])
        
        return jsonify({
            'success': True, 
            'role': row[3], 
            'username': row[1],
            'full_name': row[4],
            'requires_password_change': bool(row[6])
        })
    return jsonify({'success': False, 'error': 'Usuário ou senha inválidos'}), 401

@app.route('/api/me', methods=['GET'])
def get_current_user():
    """Verify session and return current user details."""
    if 'user_id' in session:
        return jsonify({
            'success': True,
            'username': session.get('username'),
            'role': session.get('role'),
            'full_name': session.get('full_name')
        })
    return jsonify({'success': False, 'error': 'Não logado'}), 401


@app.route('/api/admin/users', methods=['GET'])
@role_required(['master', 'admin'])
def list_users():
    print("Listing system users...")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT id, username, role, full_name, is_active FROM users')
    rows = cursor.fetchall()
    conn.close()
    
    users = []
    for r in rows:
        users.append({'id': r[0], 'username': r[1], 'role': r[2], 'full_name': r[3], 'is_active': bool(r[4])})
    print(f"Found {len(users)} users.")
    return jsonify({'success': True, 'data': users})

@app.route('/api/admin/users', methods=['POST'])
@role_required(['master', 'admin'])
def create_user():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    role = data.get('role', 'suporte')
    full_name = data.get('full_name', '').strip()
    
    print(f"Creating new user: {username} (Role: {role}, Name: {full_name})")
    
    if not username or not password or len(password) < 8:
        print("Error: Username and password (min 8 chars) are required.")
        return jsonify({'success': False, 'error': 'Usuário e senha (mín 8 caracteres) são obrigatórios'}), 400
    
    # Validar role
    valid_roles = ['suporte', 'viewer', 'admin', 'master']
    if role not in valid_roles:
        return jsonify({'success': False, 'error': 'Role inválido'}), 400
        
    # Hash da senha
    password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO users (username, password, role, full_name, is_active, requires_password_change) VALUES (?, ?, ?, ?, 1, 1)',
                       (username, password_hash, role, full_name))
        conn.commit()
        print(f"User {username} created successfully in database.")
        return jsonify({'success': True})
    except sqlite3.IntegrityError as e:
        print(f"Integrity Error: {e}")
        return jsonify({'success': False, 'error': 'Usuário já existe'}), 400
    except Exception as e:
        print(f"Unexpected error creating user: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@role_required(['master', 'admin'])
def update_user(user_id):
    data = request.get_json()
    role = data.get('role')
    full_name = data.get('full_name')
    is_active = data.get('is_active')
    password = data.get('password')
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    updates = []
    params = []
    if role:
        valid_roles = ['suporte', 'viewer', 'admin', 'master']
        if role not in valid_roles:
            return jsonify({'success': False, 'error': 'Role inválido'}), 400
        updates.append("role = ?")
        params.append(role)
    if full_name:
        updates.append("full_name = ?")
        params.append(full_name.strip())
    
    if password:
        if len(password) < 8:
            return jsonify({'success': False, 'error': 'Senha deve ter no mínimo 8 caracteres'}), 400
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        updates.append("password = ?")
        params.append(password_hash)
        updates.append("requires_password_change = 1")
        
    if is_active is not None:
        updates.append("is_active = ?")
        params.append(1 if is_active else 0)
        
    if not updates:
        return jsonify({'success': False, 'error': 'Nenhum dado para atualizar'}), 400
        
    params.append(user_id)
    set_clause = ", ".join(updates)
    cursor.execute(f'UPDATE users SET {set_clause} WHERE id = ?', params)
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@role_required('master')
def delete_user(user_id):
    if user_id == session.get('user_id'):
        return jsonify({'success': False, 'error': 'Não é possível excluir o próprio usuário'}), 400
        
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/profile', methods=['POST', 'PUT'])
def update_profile():
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Não autorizado'}), 401
        
    data = request.get_json()
    full_name = data.get('full_name', '').strip()
    password = data.get('password', '').strip()
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    updates = []
    params = []
    if full_name:
        updates.append("full_name = ?")
        params.append(full_name)
        session['full_name'] = full_name
    if password and len(password) >= 8:
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        updates.append("password = ?")
        params.append(password_hash)
        updates.append("requires_password_change = 0")
        session['requires_password_change'] = False
    elif password and len(password) < 8:
        return jsonify({'success': False, 'error': 'Senha deve ter no mínimo 8 caracteres'}), 400
        
    if not updates:
        return jsonify({'success': False, 'error': 'Nenhum dado para atualizar'}), 400
        
    params.append(session['user_id'])
    set_clause = ", ".join(updates)
    cursor.execute(f'UPDATE users SET {set_clause} WHERE id = ?', params)
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/next-laudo-num')
def next_laudo_num():
    year = datetime.now().year
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    # Find the highest number for the current year
    # id_laudo format: XXX/YYYY
    cursor.execute("SELECT id_laudo FROM laudos WHERE id_laudo LIKE ?", (f'%/{year}',))
    rows = cursor.fetchall()
    conn.close()
    
    max_num = 0
    for (id_str,) in rows:
        try:
            num = int(id_str.split('/')[0])
            if num > max_num:
                max_num = num
        except (ValueError, IndexError):
            continue
            
    next_num = f"{max_num + 1:03d}"
    return jsonify({'success': True, 'next_id': f"{next_num}/{year}"})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/tasy')
def get_tasy():
    return jsonify({'success': True, 'data': TASY_DATA_CACHE})

@app.route('/api/gerar-laudo', methods=['POST'])
@role_required(['master', 'suporte'])
def gerar_laudo():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'Dados não recebidos'}), 400
        
        # Save to DB
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO laudos (id_laudo, data, unidade, setor, local, nome_analista, descricao_problema, marca, modelo, serie, situacao, item_defeito, cargo_analista, itens_count, is_test, tipo, chamado)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('id_laudo'),
            data.get('data'),
            data.get('unidade'),
            data.get('setor'),
            data.get('local'),
            data.get('nome_analista'),
            data.get('descricao_problema'),
            data.get('marca'),
            data.get('modelo'),
            data.get('serie'),
            data.get('situacao'),
            data.get('item_defeito'),
            data.get('cargo_analista'),
            len(data.get('equipamentos', [])),
            1 if data.get('is_test') else 0,
            data.get('tipo', 'laudo'),
            data.get('chamado', '')
        ))
        conn.commit()
        conn.close()

        # Add logged in user info to data for the template
        data['preenchido_por'] = session.get('username', 'Sistema')

        from laudo_generator import generate_laudo
        output_path, temp_dir = generate_laudo(data)
        
        laudo_id = data.get('id_laudo', 'laudo').replace("/", "_")
        
        filename = os.path.basename(output_path)
        is_pdf = filename.endswith('.pdf')
        ext = '.pdf' if is_pdf else '.docx'
        
        persistent_filename = f'Laudo_{laudo_id}{ext}'
        persistent_path = os.path.join(PDF_STORAGE, persistent_filename)
        
        # Save a persistent copy locally
        shutil.copy2(output_path, persistent_path)
        
        # Attempt network copy
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute('SELECT value FROM settings WHERE key = "network_path"')
            row = cursor.fetchone()
            conn.close()
            
            network_path = row[0] if row else NETWORK_PATH_DEFAULT
            
            if network_path:
                # Debugging: Log target path
                print(f"Tentando backup para: {network_path}")
                if not os.path.exists(network_path):
                    print(f"Caminho de rede não encontrado ou inacessível: {network_path}")
                else:
                    target_file = os.path.join(network_path, persistent_filename)
                    shutil.copy2(output_path, target_file)
                    print(f"Backup concluído com sucesso: {target_file}")
        except Exception as net_err:
            print(f"Erro crítico ao copiar para rede: {net_err}")
            import traceback
            traceback.print_exc()

        # Clean up temp dir immediately
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        return jsonify({
            'success': True,
            'filename': persistent_filename,
            'is_pdf': is_pdf
        })
    
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/api/stats')
@role_required(['master', 'suporte', 'viewer', 'admin'])
def get_stats():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Total laudos
    cursor.execute('SELECT COUNT(*) FROM laudos WHERE is_test = 0 AND id_laudo NOT LIKE "IMP-%" AND tipo = "laudo"')
    total_laudos = cursor.fetchone()[0]

    # Total requests
    cursor.execute('SELECT COUNT(*) FROM laudos WHERE is_test = 0 AND id_laudo NOT LIKE "IMP-%" AND tipo = "compra"')
    total_compras = cursor.fetchone()[0]
    
    # Stats by Unit (Top 5 for chart)
    cursor.execute('SELECT unidade, COUNT(*) FROM laudos WHERE is_test = 0 AND id_laudo NOT LIKE "IMP-%" GROUP BY unidade ORDER BY COUNT(*) DESC LIMIT 5')
    unidades = cursor.fetchall()
    
    # Stats by Item (Top 5 for chart)
    cursor.execute('SELECT item_defeito, COUNT(*) FROM laudos WHERE is_test = 0 AND id_laudo NOT LIKE "IMP-%" AND item_defeito IS NOT NULL AND item_defeito != "" GROUP BY item_defeito ORDER BY COUNT(*) DESC LIMIT 5')
    itens = cursor.fetchall()
    
    # Stats by Analyst (Top 5 for chart)
    cursor.execute('SELECT nome_analista, COUNT(*) FROM laudos WHERE is_test = 0 AND id_laudo NOT LIKE "IMP-%" AND nome_analista IS NOT NULL AND nome_analista != "" GROUP BY nome_analista ORDER BY COUNT(*) DESC LIMIT 5')
    analistas = cursor.fetchall()

    # Recent laudos
    cursor.execute('SELECT id_laudo, data, unidade, setor, is_test, tipo FROM laudos WHERE id_laudo NOT LIKE "IMP-%" ORDER BY timestamp DESC LIMIT 10')
    recent = cursor.fetchall()
    
    conn.close()
    
    return jsonify({
        'success': True,
        'total': total_laudos,
        'total_compras': total_compras,
        'unidades': unidades,
        'itens': itens,
        'analistas': analistas,
        'recent': recent
    })

@app.route('/api/laudos', methods=['GET'])
@role_required(['master', 'suporte', 'viewer', 'admin'])
def get_laudos():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    tipo = request.args.get('tipo')
    if tipo:
        cursor.execute('SELECT * FROM laudos WHERE id_laudo NOT LIKE "IMP-%" AND tipo = ? ORDER BY timestamp DESC', (tipo,))
    else:
        cursor.execute('SELECT * FROM laudos WHERE id_laudo NOT LIKE "IMP-%" ORDER BY timestamp DESC')
        
    columns = [description[0] for description in cursor.description]
    rows = cursor.fetchall()
    conn.close()
    
    laudos = []
    for row in rows:
        laudos.append(dict(zip(columns, row)))
        
    return jsonify({'success': True, 'data': laudos})

@app.route('/api/laudos/<int:id>', methods=['DELETE'])
@role_required(['master', 'admin'])
def delete_laudo(id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM laudos WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/laudos/<int:id>/toggle-test', methods=['POST'])
@role_required(['master', 'admin'])
def toggle_test(id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('UPDATE laudos SET is_test = 1 - is_test WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/reports/incidences')
@role_required(['master', 'suporte', 'viewer', 'admin'])
def get_incidence_report():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Incidences of item_defeito by unidade (Gargalos)
    cursor.execute('''
        SELECT item_defeito, COUNT(*) as total_ocorr, GROUP_CONCAT(DISTINCT unidade) as unidades 
        FROM laudos 
        WHERE is_test = 0 AND id_laudo NOT LIKE "IMP-%" AND item_defeito IS NOT NULL AND item_defeito != ""
        GROUP BY item_defeito 
        ORDER BY total_ocorr DESC
    ''')
    incidences = cursor.fetchall()
    
    # Problems summary
    cursor.execute('''
        SELECT item_defeito, COUNT(*) as count 
        FROM laudos 
        WHERE is_test = 0 AND id_laudo NOT LIKE "IMP-%" AND item_defeito IS NOT NULL AND item_defeito != ""
        GROUP BY item_defeito 
        ORDER BY count DESC 
        LIMIT 10
    ''')
    problem_items = cursor.fetchall()
    
    conn.close()
    
    return jsonify({
        'success': True,
        'incidences': incidences,
        'problem_items': problem_items
    })

@app.route('/api/reports/export')
@role_required(['master', 'admin', 'suporte'])
def export_report():
    data_inicio = request.args.get('inicio', '')
    data_fim = request.args.get('fim', '')
    tipo = request.args.get('tipo', '')
    formato = request.args.get('format', 'csv')
    
    query = 'SELECT id_laudo, tipo, data, unidade, setor, nome_analista, item_defeito, situacao, descricao_problema FROM laudos WHERE is_test = 0 AND id_laudo NOT LIKE "IMP-%"'
    params = []
    
    if data_inicio:
        query += " AND data >= ?"
        params.append(data_inicio)
    if data_fim:
        query += " AND data <= ?"
        params.append(data_fim)
    if tipo:
        query += " AND tipo = ?"
        params.append(tipo)
        
    query += " ORDER BY data DESC"
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(query, params)
    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    conn.close()
    
    from flask import Response
    
    if formato == 'excel':
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatorio"
        
        ws.append(columns)
        for row in rows:
            ws.append(row)
            
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='relatorio_exportacao.xlsx'
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';', dialect='excel')
        writer.writerow(columns)
        writer.writerows(rows)
        
        response = Response(output.getvalue(), content_type='text/csv; charset=utf-8-sig')
        response.headers['Content-Disposition'] = 'attachment; filename=relatorio_exportacao.csv'
        return response

@app.route('/api/view-pdf/<path:filename>')
@role_required(['master', 'suporte', 'viewer', 'admin'])
def view_pdf(filename):
    down = request.args.get('download', '0') == '1'
    # 1. Primary storage check
    pdf_path = os.path.join(PDF_STORAGE, filename)
    if os.path.exists(pdf_path):
        mime_type = 'application/pdf' if filename.lower().endswith('.pdf') else 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        return send_file(pdf_path, mimetype=mime_type, as_attachment=down)
    
    # 2. Legacy fallback for imported records (IMP-XXX)
    # Expected filename from frontend: Laudo_IMP-035_2025.pdf
    match = re.search(r'Laudo_IMP-(\d+)', filename)
    if match:
        number_str = match.group(1)
        number_int = int(number_str)
        legacy_dir = os.path.join(BASE_DIR, 'Laudos antigos')
        
        if os.path.exists(legacy_dir):
            # Filenames in legacy folder: "Laudo 35 - ..."
            # We search for a file starting with "Laudo {number} " or "Laudo {number_with_leading_zero} "
            for f in os.listdir(legacy_dir):
                if f.lower().endswith('.pdf'):
                    # Match "Laudo 35 " or "Laudo 035 "
                    if f.startswith(f"Laudo {number_int} ") or f.startswith(f"Laudo {number_str} "):
                        return send_file(os.path.join(legacy_dir, f), mimetype='application/pdf')
    
    return "Arquivo não encontrado", 404

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                               'favicon.ico', mimetype='image/vnd.microsoft.icon')

# --- Legacy PDF Management ---

@app.route('/api/legacy-pdfs', methods=['GET'])
@role_required(['master', 'suporte', 'viewer', 'admin'])
def list_legacy_pdfs():
    """Lists all PDFs in the 'Laudos antigos' folder."""
    if not os.path.exists(LEGACY_DIR):
        return jsonify({'success': True, 'files': []})
    
    files = []
    for f in sorted(os.listdir(LEGACY_DIR)):
        if f.lower().endswith('.pdf'):
            fpath = os.path.join(LEGACY_DIR, f)
            size = os.path.getsize(fpath)
            files.append({'name': f, 'size': size})
    
    return jsonify({'success': True, 'files': files})

@app.route('/api/legacy-pdfs/<path:filename>', methods=['DELETE'])
@role_required(['master', 'admin'])
def delete_legacy_pdf(filename):
    """Permanently deletes a PDF from the 'Laudos antigos' folder."""
    # Security: ensure no path traversal
    safe_name = os.path.basename(filename)
    fpath = os.path.join(LEGACY_DIR, safe_name)
    
    if not os.path.exists(fpath):
        return jsonify({'success': False, 'error': 'Arquivo não encontrado'}), 404
    
    os.remove(fpath)
    return jsonify({'success': True})

@app.route('/api/view-legacy-pdf/<path:filename>')
@role_required(['master', 'suporte', 'viewer', 'admin'])
def view_legacy_pdf_direct(filename):
    """Directly serves a PDF from the 'Laudos antigos' folder."""
    safe_name = os.path.basename(filename)
    fpath = os.path.join(LEGACY_DIR, safe_name)
    if os.path.exists(fpath):
        return send_file(fpath, mimetype='application/pdf')
    return "Arquivo não encontrado", 404

# --- New Configuration Endpoints ---

@app.route('/api/options', methods=['GET'])
def get_options():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT id, category, value, extra FROM options')
    rows = cursor.fetchall()
    conn.close()
    
    options = []
    for row in rows:
        options.append({
            'id': row[0],
            'category': row[1],
            'value': row[2],
            'extra': json.loads(row[3]) if row[3] else None
        })
    return jsonify({'success': True, 'data': options})

@app.route('/api/options', methods=['POST'])
@role_required(['master', 'admin'])
def add_option():
    data = request.get_json()
    category = data.get('category')
    value = data.get('value')
    extra = data.get('extra')
    
    if not category or not value:
        return jsonify({'success': False, 'error': 'Categoria e valor são obrigatórios'}), 400
        
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('INSERT INTO options (category, value, extra) VALUES (?, ?, ?)',
                 (category, value, json.dumps(extra) if extra else None))
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'id': new_id})

@app.route('/api/options/<int:id>', methods=['DELETE'])
@role_required(['master', 'admin'])
def delete_option(id):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM options WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/options/<int:id>', methods=['PUT'])
@role_required(['master', 'admin'])
def update_option(id):
    data = request.get_json()
    value = data.get('value')
    extra = data.get('extra')
    
    if not value:
        return jsonify({'success': False, 'error': 'Valor é obrigatório'}), 400
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('UPDATE options SET value = ?, extra = ? WHERE id = ?',
                   (value, json.dumps(extra) if extra else None, id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/settings', methods=['GET'])
def get_settings():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT key, value FROM settings')
    rows = cursor.fetchall()
    conn.close()
    
    settings = {row[0]: row[1] for row in rows}
    return jsonify({'success': True, 'data': settings})

@app.route('/api/settings', methods=['POST'])
@role_required(['master', 'admin'])
def update_setting():
    data = request.get_json()
    key = data.get('key')
    value = data.get('value')
    
    if not key:
        return jsonify({'success': False, 'error': 'Chave é obrigatória'}), 400
        
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', (key, value))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/backup-now', methods=['POST'])
@role_required(['master', 'admin'])
def backup_now():
    """Executa backup imediato dos laudos para a rede"""
    try:
        success, message, files_count = perform_backup()
        
        if success:
            return jsonify({
                'success': True,
                'files_count': files_count,
                'message': message
            })
        else:
            return jsonify({
                'success': False,
                'message': message
            }), 400
        
    except Exception as e:
        print(f'Erro ao fazer backup: {str(e)}')
        return jsonify({
            'success': False,
            'message': f'Erro ao fazer backup: {str(e)}'
        }), 500


@app.route('/api/backup-schedule', methods=['POST', 'GET'])
@role_required(['master', 'admin'])
def backup_schedule():
    """Ativa/desativa ou retorna status do agendamento automático de backup"""
    try:
        # GET: Retornar status atual
        if request.method == 'GET':
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            cursor.execute('SELECT value FROM settings WHERE key = "auto_backup_enabled"')
            row = cursor.fetchone()
            enabled = row[0] if row else '0'
            enabled = enabled == '1'
            
            cursor.execute('SELECT value FROM settings WHERE key = "last_backup"')
            row = cursor.fetchone()
            last_backup = row[0] if row else None
            
            conn.close()
            
            # Calcular próxima execução
            from datetime import datetime, timedelta
            now = datetime.now()
            next_run = now.replace(hour=20, minute=0, second=0, microsecond=0)
            
            if next_run <= now:
                next_run += timedelta(days=1)
            
            return jsonify({
                'success': True,
                'enabled': enabled,
                'last_backup': last_backup,
                'next_run': next_run.isoformat() if enabled else None
            })
        
        # POST: Atualizar configuração
        data = request.get_json()
        enabled = data.get('enabled', False)
        
        # Atualizar database
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)',
                      ('auto_backup_enabled', '1' if enabled else '0'))
        conn.commit()
        conn.close()
        
        print(f'[BACKUP] Auto-backup {"ativado" if enabled else "desativado"}')
        
        # Reconfigurar scheduler
        setup_backup_scheduler()
        
        # Calcular próxima execução
        from datetime import datetime, timedelta
        now = datetime.now()
        next_run = now.replace(hour=20, minute=0, second=0, microsecond=0)
        
        if next_run <= now:
            next_run += timedelta(days=1)
        
        return jsonify({
            'success': True,
            'enabled': enabled,
            'next_run': next_run.isoformat() if enabled else None,
            'message': f'Agendamento de backup {"ativado" if enabled else "desativado"}'
        })
        
    except Exception as e:
        print(f'[BACKUP] Erro ao configurar agendamento: {str(e)}')
        return jsonify({
            'success': False,
            'message': f'Erro ao configurar agendamento: {str(e)}'
        }), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
